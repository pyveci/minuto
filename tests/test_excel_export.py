#!/usr/bin/env python3
"""Smoke tests for the XLSX export."""

import unittest
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory

import pytz
from openpyxl import load_workbook

from minuto.main import (
    MONTHLY_PREPAY_AMOUNT,
    CompensationPeriod,
    CompensationReport,
    CompensationType,
)


def _periods():
    """Two shifts for one user across two consecutive months."""
    user = "alice@example.com"
    base = datetime(2025, 1, 31, 17, 0, tzinfo=pytz.UTC)
    return [
        CompensationPeriod(
            user=user, start=base, end=base + timedelta(hours=14),
            hours=14.0, compensated_hours=14.0, amount=14.0 * 5.56,
            compensation_type=CompensationType.STANDARD,
        ),
        CompensationPeriod(
            user=user,
            start=datetime(2025, 2, 5, 9, 0, tzinfo=pytz.UTC),
            end=datetime(2025, 2, 5, 12, 0, tzinfo=pytz.UTC),
            hours=3.0, compensated_hours=3.0, amount=3.0 * 5.56,
            compensation_type=CompensationType.STANDARD,
            holiday_info={'name': 'Some Holiday', 'source': 'test'},
        ),
    ]


class TestExcelExport(unittest.TestCase):
    def setUp(self):
        self._tmp = TemporaryDirectory()
        self.out = Path(self._tmp.name) / 'out.xlsx'
        self.report = CompensationReport(_periods())
        self.report.export_to_excel(self.out)
        self.wb = load_workbook(self.out)  # values only; formulas stay as text

    def tearDown(self):
        self.wb.close()
        self._tmp.cleanup()

    def test_file_exists_and_is_valid_xlsx(self):
        self.assertTrue(self.out.exists())
        self.assertGreater(self.out.stat().st_size, 0)

    def test_sheets_in_expected_order(self):
        self.assertEqual(
            self.wb.sheetnames,
            ['Overview', 'Monthly Per User', 'Daily Summary', 'Detailed Shifts'],
        )

    def test_overview_shows_period_and_prepay(self):
        ws = self.wb['Overview']
        self.assertEqual(ws['A1'].value, 'On-Call Compensation Report')
        self.assertEqual(ws['A3'].value, 'Period from')
        self.assertEqual(ws['A8'].value, 'Monthly pre-pay (Pauschale)')
        self.assertEqual(ws['B8'].value, MONTHLY_PREPAY_AMOUNT)

    def test_overview_uses_live_formulas(self):
        ws = self.wb['Overview']
        # Cells in the totals block (A13:B19) are formulas referencing tables.
        formula_values = [ws.cell(row=r, column=2).value for r in range(13, 20)]
        for v in formula_values:
            self.assertIsInstance(v, str)
            self.assertTrue(v.startswith('='), f"expected formula, got {v!r}")

    def test_monthly_uses_sumifs_against_tbl_shifts(self):
        ws = self.wb['Monthly Per User']
        # Header at row 1; first data row at 2. Column E = Compensated Hours formula.
        f = ws['E2'].value
        self.assertTrue(f.startswith('=SUMIFS(tbl_shifts'), f)
        # Pre-Paid uses the named cell PrePay
        self.assertIn('PrePay', ws['G2'].value)

    def test_detailed_has_helper_columns_for_audit(self):
        ws = self.wb['Detailed Shifts']
        headers = [c.value for c in ws[1]]
        for required in ('Year-Month', 'Day', 'Is Weekend', 'Is Holiday'):
            self.assertIn(required, headers)

    def test_prepaid_total_respects_eligibility(self):
        # tbl_monthly[Pre-Paid Amount] should equal MONTHLY_PREPAY_AMOUNT * (eligible months).
        # Both periods are within the only user's data range and the user has no profile,
        # so both months are eligible → 2 * 510.
        ws = self.wb['Monthly Per User']
        eligible_col = 4  # 'Eligible'
        eligible_rows = [
            ws.cell(row=r, column=eligible_col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value  # skip totals row (no User)
        ]
        self.assertEqual(eligible_rows.count('Yes'), 2)
        self.assertEqual(eligible_rows.count('No'), 0)


if __name__ == '__main__':
    unittest.main()
